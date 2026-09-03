#!/usr/bin/env python3
"""Convert BA_Master_Tracker.xlsx into data/tracker.json consumed by the site.

Usage: python3 scripts/build_data.py path/to/BA_Master_Tracker.xlsx
Re-run this whenever the workbook changes, then commit data/tracker.json.
"""
import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

SP_BASE = "https://avepointcrm.sharepoint.com/"


def resolve(target):
    """Excel stores links relative to the workbook's own SharePoint folder.
    Strip the ../ hops and re-root them on the AvePoint tenant."""
    if not target:
        return None, True
    t = target.strip()
    if t.startswith("mailto:"):
        return t, True
    if t.startswith(("http://", "https://")):
        return t, True
    rest = re.sub(r"^(\.\./)+", "", t)
    resolved = SP_BASE + rest.lstrip("/")
    # ":x:/r/sites/..." and "sites/..." re-root cleanly; anything else is a guess.
    certain = rest.startswith(":") or rest.startswith("sites/") or rest.startswith("teams/")
    return resolved, certain


REDACT = False


def maybe_redact(value):
    """With --redact-emails, keep the mailbox hint but drop the harvestable address."""
    if not REDACT or not isinstance(value, str) or "@" not in value:
        return value
    return re.sub(r"([\w.+-]+)@([\w.-]+)", lambda m: m.group(1)[:2] + "\u2026@" + m.group(2), value)


def cell(ws, coord):
    c = ws[coord]
    v = c.value
    if isinstance(v, datetime):
        v = v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        v = v.strip()
        if v.startswith("=") or v == "#VALUE!":
            v = None
    link, certain = resolve(c.hyperlink.target if c.hyperlink else None)
    return {"text": v, "url": link, "verified": certain}


def txt(ws, coord):
    return cell(ws, coord)["text"]


def link_obj(ws, label_coord, link_coord=None, label=None):
    lc = cell(ws, link_coord or label_coord)
    name = label or txt(ws, label_coord)
    url = lc["url"]
    if not url and isinstance(lc["text"], str) and lc["text"].startswith("http"):
        url = lc["text"]
    return {"name": name, "url": url, "verified": lc["verified"]}


def build(path):
    wb = openpyxl.load_workbook(path)
    out = {
        "generatedAt": datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
        "source": Path(path).name,
        "projects": [],
        "daily": [],
        "communications": [],
    }

    # ---------------- GLASS ----------------
    ws = wb["GLASS"]
    glass_links = []
    main = cell(ws, "A2")
    if main["url"] or main["text"]:
        glass_links.append({
            "name": "GLASS master reference document",
            "url": main["url"] or main["text"],
            "group": "Key document",
            "verified": main["verified"],
        })
    for row in range(5, ws.max_row + 1):
        label = txt(ws, f"A{row}")
        if not label:
            continue
        lo = link_obj(ws, f"B{row}", label=label)
        lo["group"] = "Reference file"
        lo["note"] = txt(ws, f"B{row}") if txt(ws, f"B{row}") not in (None, "link") else None
        glass_links.append(lo)
    out["projects"].append({
        "id": "glass",
        "name": "GLASS",
        "full": "GRA GLASS (AvePoint internal)",
        "blurb": "Change requests, UAT cases and reference files for the GLASS engagement.",
        "sections": [{"title": "Reference links", "type": "links", "items": glass_links}],
    })

    # ---------------- EDRMS ADB ----------------
    ws = wb["EDRMS Tracker"]
    sites, last_sn, last_name = [], None, None
    for row in range(3, 34):
        sn = txt(ws, f"A{row}")
        name = txt(ws, f"B{row}")
        c = cell(ws, f"C{row}")
        acct = cell(ws, f"D{row}")
        if not (sn or name or c["url"] or c["text"]):
            continue
        if sn:
            last_sn, last_name = sn, name
        sites.append({
            "sn": str(sn or last_sn or ""),
            "name": name or f"{last_name} ({c['text'] or 'additional link'})",
            "label": c["text"] or "link",
            "url": c["url"],
            "verified": c["verified"],
            "account": maybe_redact((acct["text"] or "").replace("mailto:", "")) if acct["text"] else None,
        })

    # R2026.2 concerns
    ws = wb["R2026.2"]
    concerns = []
    for row in range(3, ws.max_row + 1):
        if not txt(ws, f"B{row}"):
            continue
        concerns.append({
            "sn": str(txt(ws, f"A{row}") or ""),
            "concern": txt(ws, f"B{row}"),
            "description": txt(ws, f"C{row}"),
            "source": txt(ws, f"D{row}"),
            "reason": txt(ws, f"E{row}"),
            "status": txt(ws, f"F{row}"),
        })

    # Release phases: five blocks of four columns starting at A, F, K, P, U
    ws = wb["Phases"]
    phases = []
    for base in ("A", "F", "K", "P", "U"):
        cols = [chr(ord(base) + i) for i in range(4)]
        title = txt(ws, f"{base}1")
        if not title:
            continue
        steps = []
        for row in range(3, ws.max_row + 1):
            name = txt(ws, f"{cols[1]}{row}")
            if not name:
                continue
            lc = cell(ws, f"{cols[3]}{row}")
            steps.append({
                "sn": str(txt(ws, f"{cols[0]}{row}") or ""),
                "name": name,
                "description": txt(ws, f"{cols[2]}{row}"),
                "url": lc["url"],
                "verified": lc["verified"],
            })
        phases.append({"title": title, "steps": steps})

    out["projects"].append({
        "id": "edrms",
        "name": "EDRMS ADB",
        "full": "Asian Development Bank — EDRMS / DRM implementation",
        "blurb": "Sites, backlogs, test environments and release artefacts for the ADB EDRMS programme.",
        "sections": [
            {"title": "Sites & source links", "type": "sites", "items": sites},
            {"title": "Release 2026.2 — issues & concerns", "type": "concerns", "items": concerns},
            {"title": "EDRMS release phases", "type": "phases", "items": phases},
        ],
    })

    # ---------------- LHUB ----------------
    ws = wb["LHUB Role"]
    questions = [txt(ws, f"D{r}") for r in range(3, 20)]
    questions = [q for q in questions if q and not q.lower().startswith("note:")]
    note = next((txt(ws, f"D{r}") for r in range(3, 20)
                 if (txt(ws, f"D{r}") or "").lower().startswith("note:")), None)
    steps = [txt(ws, f"F{r}") for r in range(3, ws.max_row + 1)]
    steps = [s for s in steps if s]
    refs = []
    for row in range(3, ws.max_row + 1):
        name = txt(ws, f"J{row}")
        if not name:
            continue
        lo = link_obj(ws, f"K{row}", label=name)
        lo["group"] = "Reference"
        refs.append(lo)
    for coord, name in (("A3", "FSD timeline (SharePoint)"),
                        ("A5", "LHUB JIRA creation Teams group chat"),
                        ("G4", "Sample Jira ticket (VITAELXP-22170)")):
        lo = link_obj(ws, coord, label=name)
        if lo["url"]:
            lo["group"] = "Working link"
            refs.insert(0, lo)
    out["projects"].append({
        "id": "lhub",
        "name": "LHUB",
        # Hidden from the nav and from Overview; still exported. Flip to True
        # to bring the engagement back.
        "active": False,
        "full": "LHUB — Vitae for LXP (VITAELXP) Jira intake",
        "blurb": "BA role reference: how signed-off FSDs become Jira epics, stories and sub-tasks.",
        "sections": [
            {"title": "Links & references", "type": "links", "items": refs},
            {"title": "Clarification checklist for the remote BA", "type": "list", "items": questions,
             "note": note},
            {"title": "Jira ticket creation steps", "type": "list", "items": steps},
        ],
    })

    # ---------------- Daily activity log ----------------
    # The app keeps its own activity log in the browser: a task marked done is
    # logged there, and entries can be added by hand. The workbook's 34 rows
    # were cleared on request, so nothing is imported here any more. Delete
    # this guard clause to start importing them again.
    ws = wb["Daily Tracker"]
    for row in []:
        desc = txt(ws, f"C{row}")
        if not desc:
            continue
        lc = cell(ws, f"I{row}")
        out["daily"].append({
            "sn": str(txt(ws, f"A{row}") or ""),
            "date": txt(ws, f"B{row}"),
            "task": desc,
            "category": txt(ws, f"D{row}"),
            "project": txt(ws, f"E{row}"),
            "source": txt(ws, f"F{row}"),
            "status": txt(ws, f"G{row}"),
            "remarks": txt(ws, f"H{row}"),
            "url": lc["url"],
            "verified": lc["verified"],
        })

    # ---------------- Communication tracker ----------------
    ws = wb["Communication Tracker"]
    for row in range(3, ws.max_row + 1):
        summary = txt(ws, f"E{row}")
        if not summary:
            continue
        lc = cell(ws, f"K{row}")
        out["communications"].append({
            "sn": str(txt(ws, f"A{row}") or ""),
            "date": txt(ws, f"B{row}"),
            "platform": txt(ws, f"C{row}"),
            "sender": txt(ws, f"D{row}"),
            "summary": summary,
            "area": txt(ws, f"F{row}"),
            "actionRequired": txt(ws, f"G{row}"),
            "owner": txt(ws, f"H{row}"),
            "due": txt(ws, f"I{row}"),
            "status": txt(ws, f"J{row}"),
            "url": lc["url"],
            "verified": lc["verified"],
        })

    return out


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("workbook", nargs="?", default="BA_Master_Tracker.xlsx")
    ap.add_argument("--redact-emails", action="store_true",
                    help="mask the account email addresses (use when the repository is public)")
    args = ap.parse_args()
    REDACT = args.redact_emails
    globals()["REDACT"] = args.redact_emails
    data = build(args.workbook)
    dest = Path(__file__).resolve().parent.parent / "data" / "tracker.json"
    dest.parent.mkdir(exist_ok=True)
    dest.write_text(json.dumps(data, indent=1, ensure_ascii=False))
    print(f"wrote {dest} — {len(data['projects'])} projects, "
          f"{len(data['daily'])} daily rows, {len(data['communications'])} comms rows")
